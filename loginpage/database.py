import openpyxl
import os
import hashlib

class ExcelDatabase:
    def __init__(self, file_path='user_login.xlsx'):
        self.file_path = file_path
        self.create_database_if_not_exists()

    def create_database_if_not_exists(self):
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Users'
            
            # Kolom untuk data pengguna
            sheet['A1'] = 'Username'
            sheet['B1'] = 'Password'
            sheet['C1'] = 'Role'
            
            wb.save(self.file_path)

    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()

    def register_user(self, username, password, role='kasir'):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        
        # Cek apakah username sudah ada
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                return False
        
        # Tambahkan pengguna baru
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=username)
        sheet.cell(row=next_row, column=2, value=self.hash_password(password))
        sheet.cell(row=next_row, column=3, value=role)
        
        wb.save(self.file_path)
        return True

    def validate_login(self, username, password):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        
        hashed_password = self.hash_password(password)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == hashed_password:
                return row[2]  # Kembalikan role
        
        return None
